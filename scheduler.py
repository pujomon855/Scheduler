# -*- coding: utf-8 -*-

import copy
from datetime import datetime
from itertools import combinations, permutations
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import random
import sys
import time

from filters import FILTER_PRIORITY1, FILTER_PRIORITY2, MonitorFilterManager, RemoteFilterManager
from monitors import ERole, MONITOR_ROLES_ALL, NOT_AT_OFFICE_ROLES, OUTPUT_ROLES
from monitors import assign_role_maxes, assign_remote_max, load_monitors_info

HEADER_ROW_IDX = 7
DATA_START_ROW_IDX = HEADER_ROW_IDX + 1
REMOTE_MAX_ROW_IDX = HEADER_ROW_IDX - 1
REMOTE_PER_DAY_ROW_IDX = REMOTE_MAX_ROW_IDX - 1


class ComboNotFoundException(Exception):
    """割り振りの組み合わせが見つからなかった場合に送出される例外"""
    def __init__(self, message: str):
        self.message = message


def make_schedule(excel_path):
    keep_vba = True if excel_path.endswith('xlsm') else False
    wb = openpyxl.load_workbook(excel_path, keep_vba=keep_vba)
    monitor_dict, must_work_at_office_groups = load_monitors_info(wb)

    ws = wb['latest']
    monitor_column_dict, weekday_dict = load_initial_schedules(ws, monitor_dict)
    weekdays = weekday_dict.values()
    days = len(weekdays)
    assign_role_maxes(monitor_dict, MONITOR_ROLES_ALL, days)

    filter_ws = wb['filters']
    monitor_filter_manager = MonitorFilterManager(filter_ws)
    assign_monitors(monitor_dict, weekdays, monitor_filter_manager)

    load_manual_remote_max(ws, monitor_dict, monitor_column_dict)
    max_num_of_remotes_per_day = load_remote_per_day(ws)
    assign_remote_max(monitor_dict, days, max_num_of_remotes_per_day=max_num_of_remotes_per_day)
    remote_filter_manager = RemoteFilterManager(filter_ws, must_work_at_office_groups)
    for max_num_of_remotes_per_day in range(max_num_of_remotes_per_day, 0, -1):
        cp_md, num_of_unassigned_days = assign_remotes(
            monitor_dict, sorted(weekdays), remote_filter_manager,
            max_num_of_remotes_per_day=max_num_of_remotes_per_day)
        copy_to_original_monitor_dict(cp_md, monitor_dict)
        if num_of_unassigned_days <= 0:
            break

    fill_in_blanks_to(monitor_dict, weekdays, ERole.N)
    debug_schedules(monitor_dict, weekdays)

    output_schedules(ws, monitor_dict, weekday_dict, monitor_column_dict)
    wb.save(excel_path)


def load_initial_schedules(ws: Worksheet, monitor_dict: dict):
    """
    指定シートからあらかじめ代入されている予定を読み取り、各監視者のスケジュールを初期化する。

    :param ws: ワークシート
    :param monitor_dict: 監視者の辞書(key:=name, item:=Monitor)
    :return: 監視者のlatestシートにおける列インデックスの辞書(key:=name, item:=column index),
                日付の辞書(key:=行番号, item:=datetime)
    """
    monitor_column_dict = create_monitor_col_dict(ws, monitor_dict)
    num_of_monitors = len(monitor_dict)
    weekday_dict = {}
    holiday_col = find_col_idx_by_val(ws, HEADER_ROW_IDX, 'Holiday')
    for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW_IDX, max_col=num_of_monitors + 2),
                                  DATA_START_ROW_IDX):
        day = row[0].value
        if not day:
            break
        if not is_weekday(day, ws.cell(row=row_idx, column=holiday_col)):
            continue
        weekday_dict[row_idx] = day
        for idx, monitor in enumerate(monitor_dict.values(), 1):
            val = row[idx].value
            if val:
                role = convert_val_to_role(val)
                monitor.schedule[day] = role
    return monitor_column_dict, weekday_dict


def create_monitor_col_dict(ws: Worksheet, monitor_dict: dict) -> dict:
    """
    監視者のlatestシートにおける列インデックスの辞書を作成する

    :param ws: ワークシート
    :param monitor_dict: 監視者の辞書(key:=name, item:=Monitor)
    :return: 監視者のlatestシートにおける列インデックスの辞書(key:=name, item:=column index)
    """
    monitor_column_dict = {}
    for row in ws.iter_rows(min_row=HEADER_ROW_IDX, max_row=HEADER_ROW_IDX,
                            min_col=2, max_col=len(monitor_dict) + 1):
        for cell in row:
            name = cell.value
            if name not in monitor_dict:
                raise ValueError(f'{name} is not in monitors.')
            monitor_column_dict[name] = cell.column
    return monitor_column_dict


def is_weekday(date, holiday_cell):
    if holiday_cell.value:
        return False
    return date.weekday() < 5  # 5 and 6 mean Saturday and Sunday respectively.


def convert_val_to_role(val):
    for role in ERole:
        if role.name == val:
            return role
    return ERole.OTHER


def assign_monitors(monitor_dict: dict, weekdays, filter_manager: MonitorFilterManager,
                    try_cnt1=1000, try_cnt2=1000) -> None:
    """
    監視当番の割り当てを行う。

    :param monitor_dict: 監視者の辞書(key:=name, item:=Monitor)
    :param weekdays: 営業日のIterable
    :param filter_manager: フィルタ管理クラス
    :param try_cnt1: 全フィルタを使用しての割り当て試行回数
    :param try_cnt2: 条件を緩くしての割り当て試行回数
    """
    monitors = monitor_dict.values()
    sorted_weekdays = sorted(weekdays, key=_create_weekday_sort_func(monitors))
    all_monitor_combo = list(gen_monitor_combos(monitors))
    if _try_assign_monitors(
            monitor_dict, all_monitor_combo, sorted_weekdays, filter_manager, try_cnt1, FILTER_PRIORITY2):
        return
    if _try_assign_monitors(
            monitor_dict, all_monitor_combo, sorted_weekdays, filter_manager, try_cnt2, FILTER_PRIORITY1):
        return
    _assign_monitors(monitor_dict, all_monitor_combo, sorted_weekdays, filter_manager, FILTER_PRIORITY1,
                     force_exec=True)


def gen_monitor_combos(monitors):
    """
    監視の組み合わせのgeneratorを返す。

    :param monitors: 全監視メンバー(None不可)
    :return: 監視の組み合わせ(generator)
    """
    for m1, m2, m3 in permutations(monitors, 3):
        if m1.is_fix_specialist or m2.is_fix_specialist:
            yield {ERole.AM1: m1.name, ERole.AM2: m2.name, ERole.PM: m3.name}


def _create_weekday_sort_func(monitors):
    def weekday_sort_func(weekday: datetime):
        priority = 0
        for monitor in monitors:
            role = monitor.schedule.get(weekday)
            if role:
                if role in MONITOR_ROLES_ALL:
                    priority -= 2
                else:
                    priority -= 1
        if priority >= 0:
            priority = weekday.day
        return priority

    return weekday_sort_func


def copy_monitor_dict(monitor_dict: dict):
    cp = {}
    for m, ms in monitor_dict.items():
        cp[m] = copy.copy(ms)
    return cp


def copy_to_original_monitor_dict(cp_monitor_dict: dict, org_monitor_dict: dict) -> None:
    for name, cp_monitor in cp_monitor_dict.items():
        monitor = org_monitor_dict[name]
        for day, role in cp_monitor.schedule.items():
            monitor.schedule[day] = role


def _try_assign_monitors(monitor_dict, all_monitor_combo, weekdays, fm, try_cnt, filter_priority):
    for i in range(try_cnt):
        if _assign_monitors(monitor_dict, all_monitor_combo, weekdays, fm, filter_priority):
            print(f'MONITOR: filter_priority={filter_priority}: {i + 1}: found.')
            return True
    print(f'MONITOR: filter_priority={filter_priority}: {try_cnt}: not found.')
    return False


def _assign_monitors(monitor_dict: dict, all_monitor_combo: list, weekdays, fm: MonitorFilterManager,
                     filter_priority, force_exec=False):
    """
    監視当番の割り振りを行う。

    :param monitor_dict: 監視者の辞書(key:=name, item:=Monitor)
    :param all_monitor_combo: 監視の組み合わせ(key:=ERole, item:=monitor name)のlist
    :param weekdays: 営業日のIterable
    :param fm: フィルタ管理クラス
    :param filter_priority: フィルタ優先度
    :param force_exec: 均等な割り振りが不可の場合でも、その日を除いて処理を続行する場合はTrueを設定する
    :return: 割り振りが完了した場合はTrue
    """
    # 割り振りはまずコピーに対して行う
    cp_md = copy_monitor_dict(monitor_dict)
    for day in weekdays:
        filters = fm.get_filters(cp_md.values(), day, filter_priority)
        # extract monitor combo that meets all filters.
        monitor_combos = [mc for mc in all_monitor_combo if all([f(mc) for f in filters])]
        if not monitor_combos:
            if force_exec:
                continue
            return False

        # Choice a monitor combo at random.
        monitor_combo = random.choice(monitor_combos)
        cp_md[monitor_combo[ERole.AM1]].schedule[day] = ERole.AM1
        cp_md[monitor_combo[ERole.AM2]].schedule[day] = ERole.AM2
        cp_md[monitor_combo[ERole.PM]].schedule[day] = ERole.PM

    # 割り振りが全営業日で試みられた場合のみコピーからオリジナルへ割り振りをコピーする
    copy_to_original_monitor_dict(cp_md, monitor_dict)
    return True


def load_manual_remote_max(ws: Worksheet, monitor_dict: dict, monitor_column_dict: dict):
    """
    手動で入力された在宅勤務数の上限を読み込み、MonitorScheduleに設定する。

    :param ws: 読み込むシート
    :param monitor_dict: 監視者の辞書(key:=name, item:=Monitor)
    :param monitor_column_dict: 監視者のlatestシートにおける列インデックスの辞書(key:=name, item:=column index)
    :return: None
    """
    for monitor in monitor_dict.values():
        remote_max = ws.cell(row=REMOTE_MAX_ROW_IDX, column=monitor_column_dict[monitor.name]).value
        if remote_max:
            monitor.role_max[ERole.R] = remote_max


def load_remote_per_day(ws: Worksheet) -> int:
    """
    1日の最大の在宅勤務者数を読み込む。
    入力なしの場合や負の値、数値以外が入力されている場合は0を返す。

    :param ws: 読み込むシート
    :return: 1日の最大の在宅勤務者数
    """
    remote_per_day = ws.cell(row=REMOTE_PER_DAY_ROW_IDX, column=2).value
    if isinstance(remote_per_day, int) and remote_per_day >= 0:
        return remote_per_day
    return 0


def assign_remotes(monitor_dict: dict, weekdays, filter_manager: RemoteFilterManager,
                   max_num_of_remotes_per_day=2, try_cnt1=1000, try_cnt2=10000,
                   try_cnt3=1000):
    """
    在宅勤務の割り当てを行う。
    条件によっては割り当てられない日もある。
    割り当てられない日数はtry_cnt3の試行で最も少ない日のスケジュールを採用する。

    :param monitor_dict: 監視者の辞書(key:=name, item:=Monitor)
    :param weekdays: 営業日のIterable
    :param filter_manager: フィルタ管理クラス
    :param max_num_of_remotes_per_day: 1日の在宅勤務の割り当て人数
    :param try_cnt1: 全フィルタを使用しての割り当て試行回数
    :param try_cnt2: 条件を緩くしての割り当て試行回数
    :param try_cnt3: 条件を緩くし、かつ未割当日許可での割り当て試行回数
    :return: tuple(在宅勤務を割り当てた監視スケジュールのdict, 未割当日数)
    """
    try:
        cp_md = _try_assign_remotes(monitor_dict, weekdays, filter_manager,
                                    max_num_of_remotes_per_day, try_cnt1, FILTER_PRIORITY2)
    except ComboNotFoundException as e:
        print(e.message)
    else:
        return cp_md, 0

    try:
        cp_md = _try_assign_remotes(monitor_dict, weekdays, filter_manager,
                                    max_num_of_remotes_per_day, try_cnt2, FILTER_PRIORITY1)
    except ComboNotFoundException as e:
        print(e.message)
    else:
        return cp_md, 0

    min_num_of_unassigned_days = len(weekdays)
    tmp_md = None
    for i in range(max(try_cnt3, 1)):
        cp_md, num_of_unassigned_days = _assign_remotes(
            monitor_dict, weekdays, filter_manager,
            max_num_of_remotes_per_day, FILTER_PRIORITY1, force_exec=True)
        if num_of_unassigned_days == 0:
            print(f'REMOTE2: filter_priority={FILTER_PRIORITY1}: {i + 1}: found.')
            return cp_md, 0
        if num_of_unassigned_days < min_num_of_unassigned_days:
            min_num_of_unassigned_days = num_of_unassigned_days
            tmp_md = cp_md
    print(f'Not found. max_num_of_remotes_per_day={max_num_of_remotes_per_day}.'
          f' min_num_of_unassigned_days={min_num_of_unassigned_days}')
    return tmp_md, min_num_of_unassigned_days


def _try_assign_remotes(monitor_dict: dict, weekdays, fm: RemoteFilterManager,
                        max_num_of_remotes_per_day: int, try_cnt: int, filter_priority: int):
    """
    指定回数在宅勤務の割り当てを行う。

    :param monitor_dict: monitor_dict: 監視者の辞書(key:=name, item:=Monitor)
    :param weekdays: 営業日のIterable
    :param fm: フィルタ管理クラス
    :param max_num_of_remotes_per_day: 1日の在宅勤務の割り当て人数
    :param try_cnt: 試行回数
    :param filter_priority: フィルタ優先度
    :return: 割り当てを行った監視者の辞書(コピー)
    :raises: ComboNotFoundException: 割り当てが行われなかった営業日が存在する場合
    """
    for i in range(try_cnt):
        cp_md, num_of_unassigned_days = _assign_remotes(
            monitor_dict, weekdays, fm, max_num_of_remotes_per_day, filter_priority)
        if num_of_unassigned_days == 0:
            print(f'REMOTE: filter_priority={filter_priority}, '
                  f'max_num_of_remotes_per_day={max_num_of_remotes_per_day}: {i + 1}: found.')
            return cp_md
    raise ComboNotFoundException(f'Remote combo not found. '
                                 f'filter_priority={filter_priority}, '
                                 f'max_num_of_remotes_per_day={max_num_of_remotes_per_day}: {try_cnt}')


def _assign_remotes(monitor_dict: dict, weekdays, fm: RemoteFilterManager,
                    max_num_of_remotes_per_day: int, filter_priority: int, force_exec=False):
    """
    在宅勤務の割り当てを行う。
    条件によっては割り当てられない日もある。

    :param monitor_dict: 監視者の辞書(key:=name, item:=Monitor)
    :param weekdays: 営業日のIterable
    :param fm: フィルタ管理クラス
    :param max_num_of_remotes_per_day: 1日の在宅勤務の割り当て人数
    :param filter_priority: フィルタ優先度
    :param force_exec: 均等な割り振りが不可の場合でも、その日を除いて処理を続行する場合はTrueを設定する
    :return: tuple(割り当てを行った監視者の辞書(コピー), 未割当日数)
    """
    num_of_assigned_days = 0
    # コピーに対して割り振りを行う
    cp_md = copy_monitor_dict(monitor_dict)
    monitors = cp_md.values()
    for day in weekdays:
        not_at_office_monitor_names = set()
        at_office_but_not_monitor_names = set()
        for monitor in monitors:
            role = monitor.schedule.get(day)
            if role is None:
                at_office_but_not_monitor_names.add(monitor.name)
            elif role in NOT_AT_OFFICE_ROLES:
                not_at_office_monitor_names.add(monitor.name)
        num_of_remote_monitors = max_num_of_remotes_per_day - len(not_at_office_monitor_names)
        if num_of_remote_monitors <= 0:
            num_of_assigned_days += 1
            continue

        remote_groups = []
        remote_filters = fm.get_filters(monitors, day, filter_priority)

        for group in combinations(at_office_but_not_monitor_names, num_of_remote_monitors):
            g = set(group)
            if all([f(g) for f in remote_filters]):
                remote_groups.append(g)
        if not remote_groups:
            if force_exec:
                continue
            else:
                return cp_md, len(weekdays) - num_of_assigned_days

        remote_monitor_name_set = random.choice(remote_groups)
        for monitor_name in remote_monitor_name_set:
            cp_md[monitor_name].schedule[day] = ERole.R
        num_of_assigned_days += 1

    num_of_unassigned_days = len(weekdays) - num_of_assigned_days
    if num_of_unassigned_days:
        return cp_md, num_of_unassigned_days
    return cp_md, 0


def fill_in_blanks_to(monitor_dict: dict, weekdays, role: ERole) -> None:
    for day in weekdays:
        for monitor in monitor_dict.values():
            if day not in monitor.schedule:
                monitor.schedule[day] = role


def output_schedules(ws: Worksheet, monitor_dict: dict, weekday_dict: dict,
                     monitor_column_dict: dict):
    monitor_name_st_col = find_col_idx_by_val(ws, HEADER_ROW_IDX, ERole.AM1.name)
    monitor_name_cols = {
        ERole.AM1: monitor_name_st_col,
        ERole.AM2: monitor_name_st_col + 1,
        ERole.PM: monitor_name_st_col + 2
    }
    for row_idx, weekday in weekday_dict.items():
        for monitor in monitor_dict.values():
            role = monitor.schedule.get(weekday)
            if role in OUTPUT_ROLES:
                ws.cell(row=row_idx, column=monitor_column_dict[monitor.name], value=role.name)
                col_idx = monitor_name_cols.get(role)
                if col_idx:
                    ws.cell(row=row_idx, column=col_idx, value=monitor.name)


def find_col_idx_by_val(ws: Worksheet, row_idx: int, value):
    for row in ws.iter_rows(min_row=row_idx, max_row=row_idx):
        for cell in row:
            if cell.value == value:
                return cell.column


def elapsed_time(f):
    def wrapper(*args, **kwargs):
        st = time.time()
        v = f(*args, **kwargs)
        print(f'{f.__name__}: {time.time() - st}')
        return v

    return wrapper


def debug_schedules(monitor_dict: dict, weekdays):
    date_str = 'date'
    print(f'{date_str: <19}: {ERole.AM1.name}, {ERole.AM2.name}, {ERole.PM.name}, '
          f'{ERole.N.name}, {ERole.R.name}')
    print_roles = {ERole.AM1, ERole.AM2, ERole.PM, ERole.N, ERole.R, }
    for day in sorted(weekdays):
        md = {}
        normals = []
        remotes = []
        for monitor in monitor_dict.values():
            role = monitor.schedule.get(day)
            if role in print_roles:
                if role == ERole.R:
                    remotes.append(monitor.name)
                elif role == ERole.N:
                    normals.append(monitor.name)
                else:
                    md[role] = monitor.name
        am1 = md.get(ERole.AM1)
        am2 = md.get(ERole.AM2)
        pm = md.get(ERole.PM)
        n = ' & '.join(normals) if normals else '[]'
        r = ' & '.join(remotes) if remotes else '[]'
        print(f'{day}: {am1}, {am2}, {pm}, {n}, {r}')
    print()
    print('name, AM1, AM2, PM, SUM, R')
    for name, monitor in monitor_dict.items():
        print(f'{name}, {monitor.get_role_count(ERole.AM1)}, {monitor.get_role_count(ERole.AM2)}, '
              f'{monitor.get_role_count(ERole.PM)}, {monitor.get_role_count(*MONITOR_ROLES_ALL)}, '
              f'{monitor.get_role_count(ERole.R)}')
    print()
    print('name, AM1, AM2, PM, SUM, R')
    for name, monitor in monitor_dict.items():
        print(f'{name}\', {monitor.role_max[ERole.AM1]}, {monitor.role_max[ERole.AM2]} '
              f',{monitor.role_max[ERole.PM]}, {monitor.sum_max_monitor_count}, {monitor.role_max[ERole.R]}')
    print()


@elapsed_time
def main(file_path='./schedules/MonitorSchedule2020_test.xlsm'):
    make_schedule(file_path)


if __name__ == '__main__':
    if len(sys.argv) > 1:
        main(sys.argv[1])
    else:
        main()
