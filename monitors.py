# -*- coding: utf-8 -*-

from enum import Enum, IntEnum, auto
from openpyxl.workbook import Workbook
import random


class EMonitorsColIdx(IntEnum):
    """監視者情報のNo列を基準とした相対的な列インデックス"""
    MONITOR_NO = 0
    MONITOR_NAME = 1
    FIX_SPECIALIST = 2
    MONITOR_NOTE = 3


class EComboColIdx(IntEnum):
    """組み合わせ情報のNo列を基準とした相対的な列インデックス"""
    COMBO_NO = 0
    COMBO_MEMBER1 = 1
    COMBO_MEMBER2 = 2
    COMBO_NOTE = 3


class ERole(Enum):
    """役割"""
    AM1 = auto()    # 午前当番1
    AM2 = auto()    # 午前当番2
    PM = auto()     # 午後当番
    N = auto()      # 通常
    R = auto()      # 在宅勤務
    OTHER = auto()  # 休暇・不在等


MONITOR_ROLES_ALL = {ERole.AM1, ERole.AM2, ERole.PM, }
MONITOR_ROLES_AM = {ERole.AM1, ERole.AM2, }
NOT_AT_OFFICE_ROLES = {ERole.R, ERole.OTHER, }
OUTPUT_ROLES = {r for r in ERole if r != ERole.OTHER}


class Monitor:
    """監視者情報クラス"""

    def __init__(self, name: str, is_fix_specialist: bool):
        self.name: str = name
        self.is_fix_specialist: bool = is_fix_specialist
        # 日付ごとの役割(key: datetime.datetime, item: ERole)
        self.schedule: dict = {}
        # 役割毎の最大割り当て数(key: ERole, item: max)
        self.role_max: dict = {}

    @property
    def sum_max_monitor_count(self) -> int:
        """
        :return: 監視当番の最大割り当て日数
        """
        return sum([self.role_max.get(r, 0) for r in MONITOR_ROLES_ALL])

    def is_role_max(self, role: ERole) -> bool:
        """
        この監視者に割り当てられた役割の日数が設定された上限値に達しているかを判定する。
        役割に上限が設定されていなかった場合はFalseを返す。

        :param role: 役割
        :return: この監視者に割り当てられた役割の日数が設定された上限値に達している場合はTrue
        """
        if max_count := self.role_max.get(role):
            return len([r for r in self.schedule.values() if r == role]) >= max_count
        return False

    def get_role_count(self, *roles) -> int:
        """
        指定役割に割り当てられた日数を返す。

        :param roles: 役割
        :return: 指定役割に割り当てられた日数
        """
        return len([r for r in self.schedule.values() if r in roles])

    def __repr__(self):
        return self.name

    def __copy__(self):
        cp = Monitor(self.name, self.is_fix_specialist)
        cp.schedule = self.schedule.copy()
        cp.role_max = self.role_max.copy()
        return cp


def load_monitors_info(wb: Workbook, **config):
    """
    Excelから監視者情報をよみこむ

    :param wb: workbook
    :return: [監視者の辞書]と[最低１人は出社する必要のある監視者の組み合わせのリスト]のtuple
    """
    ws = wb['monitors']
    monitor_dict = {}
    must_work_at_office_groups = []
    data_start_row_idx = config.get('DATA_START_ROW_IDX', 8)
    monitor_col_idx = config.get('MONITOR_COL_IDX', 1)
    combo_col_idx = config.get('COMBO_COL_IDX', 6)

    for row_idx in range(data_start_row_idx, ws.max_row + 1):
        # Monitors
        if name := ws.cell(row_idx, monitor_col_idx + EMonitorsColIdx.MONITOR_NAME).value:
            fix_specialist_val = ws.cell(
                row_idx, monitor_col_idx + EMonitorsColIdx.FIX_SPECIALIST).value
            monitor_dict[name] = Monitor(name, fix_specialist_val == 1)

        # Groups
        if name1 := ws.cell(row_idx, combo_col_idx + EComboColIdx.COMBO_MEMBER1).value:
            name2 = ws.cell(row_idx, combo_col_idx + EComboColIdx.COMBO_MEMBER2).value
            must_work_at_office_groups.append({name1, name2})
    return monitor_dict, must_work_at_office_groups


def assign_role_maxes(monitor_dict: dict, roles, days: int) -> None:
    """
    各監視者に割り当てられた役割の日数の上限値の合計が等しくなるようにランダムに上限を設定する。
    合計値の差は最大1とする。

    :param monitor_dict: 上限を設定するMonitorの辞書(key:=MonitorName, Item:=Monitor)
    :param roles: ERoleのIterable
    :param days: 割り当て日数
    :return: None
    """
    num_of_monitors = len(monitor_dict)
    min_max_cnt = int(days / num_of_monitors)  # 最小の最大割り当て日数
    num_of_hi_freq_monitors = days % num_of_monitors
    monitors = monitor_dict.values()
    if num_of_hi_freq_monitors == 0:
        for monitor in monitors:
            for role in roles:
                monitor.role_max[role] = min_max_cnt
        return

    monitor_names = monitor_dict.keys()
    hi_freq_monitor_names = set(random.sample(monitor_names, num_of_hi_freq_monitors))
    for role in roles:
        for monitor_name in hi_freq_monitor_names:
            monitor_dict[monitor_name].role_max[role] = min_max_cnt + 1

        lo_freq_monitor_names = monitor_names - hi_freq_monitor_names
        for monitor_name in lo_freq_monitor_names:
            monitor_dict[monitor_name].role_max[role] = min_max_cnt
        hi_freq_monitor_names = _find_lower_frequency(monitors, num_of_hi_freq_monitors)


def _find_lower_frequency(monitors, find_num):
    """
    監視当番数の合計が少ない監視者のsetを、サイズがfind_numとなるように返す。
    合計が同値でfind_numを超える監視者は、同値の中でランダムに選出される。
    例：監視当番の合計が[4, 4, 5, 5, 5, 6]で、find_numが3の場合、
        返されるsetは監視当番の合計が4のMonitorと、5のMonitorの中からランダムに1人のsetとなる。

    :param monitors: MonitorのIterable
    :param find_num: 検索する監視者数(0 < find_num <= len(monitor_schedule_dict))
    :return: 監視当番数の合計が少ない監視者名のset
    """
    sorted_monitors = sorted(monitors, key=_monitor_sort_func)
    monitor_name_set = set()
    cur_mon_cnt = 0
    for monitor in sorted_monitors:
        if find_num <= len(monitor_name_set) and cur_mon_cnt < monitor.sum_max_monitor_count:
            break
        monitor_name_set.add(monitor.name)
        cur_mon_cnt = monitor.sum_max_monitor_count

    if find_num >= len(monitor_name_set):
        return monitor_name_set

    # extract monitors whose monitor_count is lower than cur_mon_cnt
    tmp_monitor_name_set = {
        monitor.name for monitor in sorted_monitors if monitor.sum_max_monitor_count < cur_mon_cnt}
    # now, monitor_set is only including monitors whose monitor_count is cur_mon_cnt
    monitor_name_set -= tmp_monitor_name_set
    # returns tmp_monitor_set and selected monitors from monitor_set at random
    monitor_name_set = tmp_monitor_name_set | set(
        random.sample(monitor_name_set, find_num - len(tmp_monitor_name_set)))
    return monitor_name_set


def _monitor_sort_func(monitor: Monitor):
    return monitor.sum_max_monitor_count


def assign_remote_max(monitor_dict: dict, days: int, max_num_of_remotes_per_day: int = 2) -> None:
    """
    在宅勤務日数の上限を均等に割り振る。
    手動での在宅勤務日数の最大値の読み取りや、休暇や不在の予定の読み取りは完了後に呼び出されることを前提としている。

    :param monitor_dict: 在宅勤務日数の上限を設定するMonitorの辞書(key:=MonitorName, Item:=Monitor)
    :param days: 割り当て日数
    :param max_num_of_remotes_per_day: 1日の在宅勤務者の最大人数(default=2)
    :return: None
    """
    manually_assigned_monitors = []
    not_manually_assigned_monitors = []
    manual_remote_max = 0
    not_work_at_office_days = 0
    for monitor in monitor_dict.values():
        if remote_max := monitor.role_max.get(ERole.R):
            manually_assigned_monitors.append(monitor)
            manual_remote_max += remote_max
        else:
            not_manually_assigned_monitors.append(monitor)
        not_work_at_office_days += monitor.get_role_count(*NOT_AT_OFFICE_ROLES)
    rem_remote_days = days * max_num_of_remotes_per_day - manual_remote_max

    if rem_remote_days <= 0:
        _set_remote_max(not_manually_assigned_monitors, 0)
        return

    num_of_not_manually_assigned_monitors = len(not_manually_assigned_monitors)
    min_remote_max = int(rem_remote_days / num_of_not_manually_assigned_monitors)
    num_of_hi_freq_monitors = rem_remote_days % num_of_not_manually_assigned_monitors
    if num_of_hi_freq_monitors == 0:
        _set_remote_max(not_manually_assigned_monitors, min_remote_max)
        return

    hi_freq_ms_list = random.sample(not_manually_assigned_monitors, num_of_hi_freq_monitors)
    _set_remote_max(hi_freq_ms_list, min_remote_max + 1)
    lo_freq_ms_list = [monitor for monitor in not_manually_assigned_monitors
                       if monitor not in hi_freq_ms_list]
    _set_remote_max(lo_freq_ms_list, min_remote_max)


def _set_remote_max(monitors, remote_max: int):
    for monitor in monitors:
        monitor.role_max[ERole.R] = max(remote_max - monitor.get_role_count(ERole.OTHER), 0)
