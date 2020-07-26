# -*- coding: utf-8 -*-

from datetime import timedelta
import math
import openpyxl
import random

from combo import ERole, MonitorSchedule, gen_monitor_combos
import monitors

HEADER_ROW_IDX = 7
DATA_START_ROW_IDX = HEADER_ROW_IDX + 1
MONITOR_ROLES_ALL = {ERole.AM1, ERole.AM2, ERole.PM, }
MONITOR_ROLES_AM = {ERole.AM1, ERole.AM2, }


def make_schedule(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    monitor_dict, must_work_at_office_groups = monitors.load_monitors_info(wb)
    ws = wb['latest']

    is_assigned = False
    for i in range(100):
        monitor_schedule_dict, weekdays = load_initial_schedules(ws, monitor_dict)
        if assign_monitors(monitor_schedule_dict, weekdays):
            is_assigned = True
            print(f'{i + 1}: found.')
            break
    if not is_assigned:
        print('could not assigned.')
        monitor_schedule_dict, weekdays = load_initial_schedules(ws, monitor_dict)
        assign_monitors(monitor_schedule_dict, weekdays, True)

    # for mc in gen_monitor_combos(monitor_dict.values()):
    #     print(f'{mc.monitor_am1.name},{mc.monitor_am2.name},{mc.monitor_pm.name}')

    for day in weekdays:
        md = {}
        for ms in monitor_schedule_dict.values():
            if (role := ms.schedule.get(day)) in MONITOR_ROLES_ALL:
                md[role] = ms.monitor
        if len(md) < 3:
            print(day)
        else:
            print(f'{day}: {md[ERole.AM1].name}, {md[ERole.AM2].name}, {md[ERole.PM].name}')


def load_initial_schedules(ws, monitor_dict):
    monitor_schedule_dict = init_monitor_schedules(ws, monitor_dict)
    num_of_monitors = len(monitor_dict)
    monitor_list = monitor_dict.values()
    weekdays = []
    for row in ws.iter_rows(min_row=DATA_START_ROW_IDX, max_col=num_of_monitors + 2):
        day = row[0].value
        if not day:
            break
        holiday_cell = row[num_of_monitors + 1]
        if not is_weekday(day, holiday_cell):
            continue
        weekdays.append(day)
        for idx, monitor in enumerate(monitor_list, 1):
            if val := row[idx].value:
                schedule = monitor_schedule_dict[monitor]
                schedule.schedule[day] = convert_val_to_role(val)
    return monitor_schedule_dict, weekdays


def init_monitor_schedules(ws, monitor_dict):
    monitor_schedule_dict = {}
    for row in ws.iter_rows(min_row=HEADER_ROW_IDX, max_row=HEADER_ROW_IDX,
                            min_col=2, max_col=len(monitor_dict) + 1):
        for cell in row:
            name = cell.value
            if name not in monitor_dict:
                raise ValueError(f'{name} is not in monitors.')
            monitor = monitor_dict[name]
            monitor_schedule_dict[monitor] = MonitorSchedule(monitor, cell.column)
    return monitor_schedule_dict


def is_weekday(date, holiday_cell):
    if holiday_cell.value:
        return False
    return date.weekday() < 5  # 5 and 6 mean Saturday and Sunday respectively.


def convert_val_to_role(val):
    for role in ERole:
        if role.name == val:
            return role
    return ERole.OTHER


def assign_monitors(monitor_schedule_dict, weekdays, force_exec=False):
    monitor_set = monitor_schedule_dict.keys()
    largest_monitoring_days = math.ceil(len(weekdays) / len(monitor_set))
    all_monitor_combos = set(gen_monitor_combos(monitor_set))
    for day in weekdays:
        filters = create_filters(day, monitor_schedule_dict, largest_monitoring_days)
        # extract monitor combo that meets all filters.
        monitor_combos = [mc for mc in all_monitor_combos if all([f(mc) for f in filters])]
        if not monitor_combos:
            print(f'{day}: no valid combo.')
            filters = create_filters(day, monitor_schedule_dict, largest_monitoring_days, False)
            monitor_combos = [mc for mc in all_monitor_combos if all([f(mc) for f in filters])]
            if not monitor_combos:
                # monitor_combos = all_monitor_combos
                if not force_exec:
                    return False

        # Choice a monitor combo at random.
        monitor_combo = random.choice(monitor_combos)
        monitor_schedule_dict[monitor_combo.monitor_am1].schedule[day] = ERole.AM1
        monitor_schedule_dict[monitor_combo.monitor_am2].schedule[day] = ERole.AM2
        monitor_schedule_dict[monitor_combo.monitor_pm].schedule[day] = ERole.PM
    return True


def create_filters(day, monitor_schedule_dict, largest_monitoring_days,
                   include_pre_day=True):
    filters = []
    for ms in monitor_schedule_dict.values():
        # Manually input day role
        if role := ms.schedule.get(day):
            if role in MONITOR_ROLES_ALL:
                filters.append(create_filter(ms.monitor, roles=[role]))
            elif role == ERole.OTHER:
                filters.append(create_filter(ms.monitor, include=False))
        else:
            # Adjust monitoring days
            filter_roles = []
            if ms.am1_count >= largest_monitoring_days:
                filter_roles.append(ERole.AM1)
            if ms.am2_count >= largest_monitoring_days:
                filter_roles.append(ERole.AM2)
            if ms.pm_count >= largest_monitoring_days:
                filter_roles.append(ERole.PM)
            if filter_roles:
                filters.append(create_filter(ms.monitor, include=False, roles=filter_roles))

        # Not to monitor am if monitor pm 1 day before.
        if include_pre_day:
            pre_day = day - timedelta(days=1)
            if role := ms.schedule.get(pre_day):
                if role == ERole.PM:
                    filters.append(create_filter(ms.monitor, include=False, roles=MONITOR_ROLES_AM))
    return filters


def create_filter(monitor, include=True, roles=None):
    if include:
        return lambda monitor_combo: monitor_combo.contains_monitor(monitor, roles)
    else:
        return lambda monitor_combo: not monitor_combo.contains_monitor(monitor, roles)


def main():
    make_schedule('./schedules/MonitorSchedule2020.xlsm')


if __name__ == '__main__':
    main()
